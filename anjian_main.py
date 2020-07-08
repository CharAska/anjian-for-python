import keyboard
import mouse
import time
import xlrd
#import xlwt

#import pandas as pd
from tkinter import *
from tkinter import messagebox #消息框模块
import tkinter.filedialog #文件导入模块
#import re #正则表达式
#import os
#import openpyxl
import sys
import os
#打包命令
#pyinstaller -F --hidden-import pandas --hidden-import="pandas._libs.tslibs.timedeltas" --hidden-import="pandas._libs.tslibs.np_datetime" --hidden-import="pandas._libs.tslibs.nattype" --hidden-import="pandas._libs.skiplist" anjian_main.py -p anjian模板.xlsx
#pyinstaller -F anjian_main.py -p anjian_temp.xlsx
#pyi-makespec -F -w anjian_main.py
#pyinstaller -F -w anjian_main.spec
#pandas要求降级0.20.3

#keyboard.wait("f3")
#keyboard.send("ctrl+c")
#keyboard.write("测试abc")
#mouse.move(1700,550)
#mouse.click()
#time.sleep(0.5)

func_dict = {"send": keyboard.send, "write": keyboard.write, "move": mouse.move, "click": mouse.click, "rightclick": mouse.right_click, "delay": time.sleep, "doubleclick":mouse.double_click}
def func_None():
    print ("cannot find func")
def anjian(i,x,y,z):
    if i == "move":
        return func_dict.get(i, func_None)(x,y)
    elif i == "click" or i == "rightclick" or i == "doubleclick":
        return func_dict.get(i, func_None)()
    else:
        return func_dict.get(i, func_None)(x)

#生成资源文件目录访问路径
def resource_path(relative_path):
    if getattr(sys, 'frozen', False): #是否Bundle Resource
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def downloadFile():
    from openpyxl import load_workbook
    tempfile = resource_path(os.path.join("res", "anjian_temp.xlsx"))
    workbook = load_workbook(tempfile)
    #workbook = load_workbook(r'./anjian_temp.xlsx')
    filename = tkinter.filedialog.asksaveasfilename(filetypes=[('xlsx', '*.xlsx')])
    filename = filename + '.xlsx'
    workbook.save(filename)
    print("download complete")

def chooseFile():
    global data
    exlFile = tkinter.filedialog.askopenfilename()
    data = xlrd.open_workbook(exlFile)
    #label_fileName.config(text=re.sub(r".*\/", "", exlFile, count=0))

global loadState
loadState = 0
def exlRead():
    # 完成后需要修改地址，可改成导入方式
    # 曲线表
    global loadState
    global action_rows, action_cols, table_action, data
    #data = xlrd.open_workbook('D:\CODE\Python\otvReach\otvReach.xlsx')  # 获取工作表
    table_action = data.sheet_by_name(u'action')  # 获取sheet
    action_rows = table_action.nrows  # 获取行数
    #action_cols = table_action.ncols  # 获取列数,可以固定为4
    action_cols = 4
    print(action_rows)
    print(action_cols)
    print("excel读取完成")
    loadState = 1

def actionList_exe():
    # 获取生成action的List
    global actionList
    actionList = []
    actionList.clear()#重置动作列表
    for i in range(1,action_rows):
        row = []
        for j in range(action_cols):
            row.append(table_action.cell(i, j).value)
        actionList.append(row)
    print(actionList)

def openFile():
    chooseFile()
    exlRead()
    actionList_exe()
    label_state1.config(text="列表加载完成")
    #messagebox.showinfo(message="导入完成")

def action_exe(x):
    print(actionList)
    for i in range(0, action_rows-1):
        anjian(actionList[i][0],actionList[i][1],actionList[i][2],actionList[i][3])
    print("Done")

def start_exe():
    global actionList
    global loadState
    global runState
    if loadState == 0:
        messagebox.showinfo("提示","请上传动作列表文件")
    elif entry_hotkey.get() == "":
        messagebox.showinfo("提示", "请输入启动热键")
    else:
        if runState == 0:
            keyboard.on_release_key(entry_hotkey.get(), action_exe, suppress=False)
            label_state2.config(text="运行中")
            button_run.config(text="停止")
            runState = 1
            print("run")
        elif runState == 1:
            keyboard.unhook_all()
            label_state2.config(text="已停止")
            button_run.config(text="运行")
            runState = 0
            print("stop")

# GUI界面代码
ui_top = Tk()
#ui_top.title('OTV项目Reach预估工具 Ver0.1')
# 边框
label_top = Label(ui_top, text="", width=3)
label_top.grid(row=0, column=0)
#label_sign = Label(ui_top, text="Ver0.1 by Char", height=3, width=20, anchor="sw")
#label_sign.grid(row=8, column=1)
label_end = Label(ui_top, text="", width=3)
label_end.grid(row=10, column=6)

# 主体
button_chooseFile = Button(ui_top, text="下载动作模板",height=2, width=15, command=lambda: downloadFile())
button_chooseFile.grid(row=1, column=1)
button_chooseFile = Button(ui_top, text="上传动作列表",height=2, width=15, command=lambda: openFile())
button_chooseFile.grid(row=2, column=1)
label_state1 = Label(ui_top, text="未上传动作列表文件", height=4, width=15)
label_state1.grid(row=2, column=2)
#label_fileName = Label(ui_top, text="尚未上传", height=3, width=20)
#label_fileName.grid(row=2, column=2)
label_hotkey = Label(ui_top, text="设置热键", height=4, width=15)
label_hotkey.grid(row=3, column=1)
global runHotkey, runState
runHotkey = Variable()
runState = 0
runHotkey.set("f3")
entry_hotkey = Entry(ui_top, textvariable=runHotkey, width=15)
entry_hotkey.grid(row=3, column=2)
button_run = Button(ui_top, text="运行", height=2,width=15, command=lambda: start_exe())
button_run.grid(row=4, column=1)
label_state2 = Label(ui_top, text="未运行", height=4, width=15)
label_state2.grid(row=4, column=2)

ui_top.title('按键精灵python版')




label_sign = Label(ui_top, text="Ver1.0 by Char", height=2, width=15, anchor="se")
label_sign.grid(row=10, column=2)
ui_top.mainloop()





